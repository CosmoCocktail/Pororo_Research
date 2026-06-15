import streamlit as st
import pandas as pd
import re
import os
import random
import base64
from openpyxl import load_workbook, Workbook

# ======================
# 페이지 설정
# ======================
st.set_page_config(
    page_title="팀플 빌런즈 : 눈 속 마을 빌런 테스트",
    page_icon="🐧",
    layout="centered"
)

# ======================
# 상수
# ======================
QUESTION_FILE   = "questions.xlsx"
RESULT_FILE     = "result.xlsx"
CHARACTERS      = ["에디", "크롱", "뽀로로", "루피", "포비"]
BG_IMAGE_FILE   = "background.png"
FONT_IMPORT_URL = "https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;700;900&display=swap"
FONT_FAMILY     = "'Noto Sans KR', sans-serif"

# ======================
# 동점 추가 질문
# ======================
TIEBREAK_QUESTION = {
    "질문":  "주제를 정하는데 의견이 갈린다면?",
    "에디":   '"잠깐, 지금 의견 말고 나 완전 좋은 생각 났어."',
    "크롱":   '"다 괜찮은 것 같은데…"',
    "뽀로로": '"제일 재밌어 보이는 거 하면 안 돼?"',
    "루피":   '"일단 기준부터 정하고 제일 괜찮은 안으로 가자."',
    "포비":   '"다들 말해봐. 내가 의견을 정리해볼게."',
}

# ======================
# 데이터 로드
# ======================
@st.cache_data
def load_questions():
    df = pd.read_excel(QUESTION_FILE)
    df = df.dropna(subset=["질문", "답변A", "답변B"])
    df = df[df["질문"].astype(str).str.strip() != ""]
    df = df[df["답변A"].astype(str).str.strip().str.lower() != "nan"]
    df = df[df["답변B"].astype(str).str.strip().str.lower() != "nan"]
    return df.reset_index(drop=True)

questions_df = load_questions()

# ======================
# 유틸 함수
# ======================
def parse_types(type_str: str) -> list:
    return [c for c in re.findall(r'[\w가-힣]+', str(type_str)) if c in CHARACTERS]

def safe_score(val) -> int:
    return int(float(val)) if str(val) not in ("", "nan") else 2

def show_image(path: str, columns=(1, 3, 1)):
    if os.path.exists(path):
        try:
            col_l, col_c, col_r = st.columns(columns)
            with col_c:
                st.image(path, use_container_width=True)
        except Exception:
            pass

# ======================
# result.xlsx — 저장 / 읽기
# ======================
def save_result(character: str):
    if os.path.exists(RESULT_FILE):
        wb = load_workbook(RESULT_FILE)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        if character in header:
            col_idx = header.index(character) + 1
            ws.cell(row=2, column=col_idx).value = (ws.cell(row=2, column=col_idx).value or 0) + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(CHARACTERS)
        ws.append([1 if c == character else 0 for c in CHARACTERS])
    wb.save(RESULT_FILE)

def load_result_counts():
    if not os.path.exists(RESULT_FILE):
        return {c: 0 for c in CHARACTERS}, 0
    df = pd.read_excel(RESULT_FILE, header=0)
    if not all(c in df.columns for c in CHARACTERS):
        return {c: 0 for c in CHARACTERS}, 0
    counts = {c: int(df[c].iloc[0]) for c in CHARACTERS}
    return counts, sum(counts.values())

# ======================
# 라우팅
# ======================
def evaluate_and_route():
    scores    = st.session_state.scores
    max_score = max(scores.values())
    top       = [c for c, s in scores.items() if s == max_score]
    if len(top) == 1:
        finalize(top[0])
    else:
        st.session_state.tied_chars = top
        st.session_state.tb_order   = None
        st.session_state.page       = "tiebreak"
    st.rerun()

def finalize(character: str):
    st.session_state.result_character = character
    if not st.session_state.result_saved:
        save_result(character)
        st.session_state.result_saved = True
    st.session_state.page = "result"

# ======================
# 세션
# ======================
DEFAULTS = {
    "page":             "home",
    "question_idx":     0,
    "scores":           {c: 0 for c in CHARACTERS},
    "result_character": None,
    "result_saved":     False,
    "tied_chars":       [],
    "tb_order":         None,
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

def reset_session():
    for k, v in DEFAULTS.items():
        st.session_state[k] = v if k != "scores" else {c: 0 for c in CHARACTERS}
    for k in [k for k in st.session_state if k.startswith("q_order_")]:
        del st.session_state[k]

# ======================
# 스타일 — 배경 / 폰트
# ======================
if os.path.exists(BG_IMAGE_FILE):
    bg_b64 = base64.b64encode(open(BG_IMAGE_FILE, "rb").read()).decode()
    bg_css = (
        "<style>.stApp {"
        + "background-image: url('data:image/png;base64," + bg_b64 + "');"
        + "background-size:cover; background-position:center;"
        + "background-repeat:no-repeat; background-attachment:fixed;"
        + "}</style>"
    )
    st.markdown(bg_css, unsafe_allow_html=True)

_font_import = ("@import url('" + FONT_IMPORT_URL + "');") if FONT_IMPORT_URL else ""

_css_parts = [
    "<style>",
    _font_import,
    "html, body, [class*='css'], .stMarkdown, .stButton button {",
    "    font-family: " + FONT_FAMILY + ";",
    "}",
    ".big-title  { text-align:center; font-size:2rem; font-weight:800; margin-bottom:0.5rem; }",
    ".sub-title  { text-align:center; font-size:1.1rem; color:#666; margin-bottom:1.5rem; }",
    ".res-char   { text-align:center; font-size:2.4rem; font-weight:900; margin:1rem 0; }",
    ".stat-box   { background:#f0f4ff; border-radius:12px; padding:1.2rem; margin-top:1rem; text-align:center; }",
    ".stat-pct   { font-size:1.2rem; margin-top:0.5rem; }",
    ".q-text     { text-align:center; font-size:1.35rem; font-weight:700; margin-bottom:1.2rem; line-height:1.5; color:#1A7FD4; }",
    ".prog-text  { text-align:center; font-size:0.9rem; color:#1A7FD4; margin-bottom:0.4rem; }",
    ".stProgress > div > div > div > div { background-color: #1A7FD4; }",
    "div[data-testid='stHorizontalBlock'] .stButton button { background-color:#1A7FD4; color:#fff; border:none; border-radius:10px; font-weight:700; }",
    "div[data-testid='stHorizontalBlock'] .stButton button:hover { background-color:#155FA0; color:#fff; }",
    ".tie-badge  { display:inline-block; background:#1A7FD4; color:#fff; border-radius:20px; padding:4px 18px; font-size:0.9rem; margin-bottom:1rem; }",
    "</style>",
]
st.markdown("\n".join(_css_parts), unsafe_allow_html=True)


# ══════════════════════════════════════════
# 홈 페이지
# ══════════════════════════════════════════
if st.session_state.page == "home":

    st.markdown("<div class='big-title'>팀플 빌런즈</div>", unsafe_allow_html=True)
    st.markdown("<div class='sub-title'>눈 속 마을 빌런 테스트</div>", unsafe_allow_html=True)
    show_image("단체컷.png")
    st.write("")

    if st.button("🐧 시작하기", use_container_width=True):
        reset_session()
        st.session_state.page = "question"
        st.rerun()


# ══════════════════════════════════════════
# 질문 페이지
# ══════════════════════════════════════════
elif st.session_state.page == "question":

    total = len(questions_df)
    idx   = st.session_state.question_idx

    if idx >= total:
        evaluate_and_route()
        st.stop()

    row = questions_df.iloc[idx]

    # 진행 상황
    st.markdown(f"<div class='prog-text'>질문 {idx+1} / {total}</div>", unsafe_allow_html=True)
    st.progress(idx / total)
    st.write("")

    # 질문 텍스트
    st.markdown(f"<div class='q-text'>{row['질문']}</div>", unsafe_allow_html=True)

    # 질문 이미지
    show_image(f"질문 이미지/question.{idx+1}.png")
    st.write("")

    # A/B 답변 (최초 진입 시 랜덤 셔플, 이후 순서 고정)
    order_key = f"q_order_{idx}"
    if order_key not in st.session_state:
        opts = [
            ("A", str(row["답변A"]), str(row["답변A_유형"]), safe_score(row["답변A_점수"])),
            ("B", str(row["답변B"]), str(row["답변B_유형"]), safe_score(row["답변B_점수"])),
        ]
        random.shuffle(opts)
        st.session_state[order_key] = opts

    col1, col2 = st.columns(2)
    for col, (label, text, types_str, score) in zip([col1, col2], st.session_state[order_key]):
        with col:
            if not text or text.strip() in ("", "nan"):
                continue
            if st.button(text, use_container_width=True, key=f"q{idx}_{label}"):
                for char in parse_types(types_str):
                    st.session_state.scores[char] += score
                st.session_state.question_idx += 1
                st.rerun()


# ══════════════════════════════════════════
# 동점 추가 질문 페이지
# ══════════════════════════════════════════
elif st.session_state.page == "tiebreak":

    tied       = st.session_state.tied_chars
    valid_tied = [c for c in tied if c in TIEBREAK_QUESTION]

    if len(valid_tied) <= 1:
        finalize(valid_tied[0] if valid_tied else tied[0])
        st.rerun()

    st.markdown(
        "<div style='text-align:center'>"
        "<span class='tie-badge'>⚖️ 두구두구! 마지막 질문</span>"
        "</div>",
        unsafe_allow_html=True
    )
    st.write("")
    st.markdown(f"<div class='q-text'>{TIEBREAK_QUESTION['질문']}</div>", unsafe_allow_html=True)
    st.write("")

    if st.session_state.tb_order is None:
        opts = [(c, TIEBREAK_QUESTION[c]) for c in valid_tied]
        random.shuffle(opts)
        st.session_state.tb_order = opts

    cols = st.columns(len(st.session_state.tb_order))
    for col, (char, text) in zip(cols, st.session_state.tb_order):
        with col:
            if st.button(text, use_container_width=True, key=f"tb_{char}"):
                finalize(char)
                st.rerun()


# ══════════════════════════════════════════
# 결과 페이지
# ══════════════════════════════════════════
elif st.session_state.page == "result":

    if not st.session_state.result_character:
        st.session_state.page = "home"
        st.rerun()

    character     = st.session_state.result_character
    counts, total = load_result_counts()

    st.markdown("<div class='big-title'>🎉 결과 발표!</div>", unsafe_allow_html=True)
    st.write("")

    # 결과 캐릭터 이미지
    img_path = f"result_page/{character}.png"
    if os.path.exists(img_path):
        show_image(img_path)
    else:
        st.info(f"📁 이미지 파일 미등록: result_page/{character}.png")

    st.markdown(
        f"<div class='res-char'>당신은 <span style='color:#1A7FD4'>{character}</span>입니다!</div>",
        unsafe_allow_html=True
    )
    st.write("")

    # 참여 통계
    count = counts.get(character, 0)
    pct   = round(count / total * 100, 1) if total > 0 else 0.0
    st.markdown(f"""
    <div class='stat-box'>
        <div style='font-size:1rem;color:#444;margin-bottom:0.4rem'>지금까지 이 테스트에 참여한 인원</div>
        <div class='stat-pct'><b>{pct}%</b> 의 인원이 <b>{character}</b> 를 선택했습니다.</div>
    </div>
    """, unsafe_allow_html=True)
    st.write("")

    if st.button("🔄 다시하기", use_container_width=True):
        reset_session()
        st.rerun()
